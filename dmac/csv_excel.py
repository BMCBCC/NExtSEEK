import csv
import openpyxl
from openpyxl.cell.read_only import EmptyCell
from dmac.conversion import toString

def handle_uploaded_file(file, output):
    destination = open(output, 'wb+')
    for chunk in file.chunks():
        destination.write(chunk)
    destination.close()    
         
def csv_dict_list(incsvfile):
    reader = csv.DictReader(open(incsvfile, 'rb'))
    dict_list = []
    for line in reader:
        dict_list.append(line)
    return dict_list

def load_csvfile(csvfile):
    filename = csvfile.name
    csvfiledata = {}
    if csvfile:
        csvfiledata['file'] = csvfile
        reader = csv.reader(csvfile)
        try:
            columns = reader.next()
            csvfiledata['columns'] = columns
            rows = []
            n = 0
            for row in reader:
                col0 = row[0]
                if len(col0)>0:
                    rows.append(row)
                n += 1
            
            msg = "Total rows loaded from input csv file: " + str(len(rows))
            csvfiledata['rows'] = rows
            csvfiledata['status'] = 1
            csvfiledata['msg'] = msg
        except csv.Error as e:
            msg = "Error in reading csv file: " + str(filename)
            csvfiledata['status'] = 0
            csvfiledata['msg'] = msg
        
    return csvfiledata     
      

def convertToDic(columns, row):
    datadic = {}
    if len(columns)!=len(row):
        return datadic
    for index, item in enumerate(row):
        key = columns[index]
        datadic[key] = item
        
    return datadic
      
def convertToDicList(columns, rows):
    datadiclist = []
    for row in rows:
        datadic = convertToDic(columns, row)
        datadiclist.append(datadic)
    
    return datadiclist
      
def checkHeaders(headersExpected, headersUploaded):
    headersExpectedNew = [x.lower() for x in headersExpected]
    headersUploadedNew = [x.lower() for x in headersUploaded]
    
    n = 0
    missing = []
    for header in headersExpectedNew:
        if header not in headersUploadedNew:
            n += 1
            missing.append(header)
                
    return n, missing 
      
def load_csvfile_diclist(csvfile, headersKnown=[]):    
    csvdata = load_csvfile(csvfile)
    csv_diclist = []
    if csvdata['status']:
        logger.debug('csvfile loaded okay')
    else:
        return csv_diclist
    
    columns = csvdata['columns']
    if len(headersKnown)>0:
        n, missing = checkHeaders(headersKnown, columns)
        if n>0:
            msg = "Miss the following columns: " + ','.join(missing)
            csvdata['status'] = 0
            csvdata['msg'] = msg
            csvdata['diclist'] = csv_diclist
            return csvdata
        
    rows = csvdata['rows']
    csv_diclist = convertToDicList(columns, rows)
    csvdata['diclist'] = csv_diclist
    csvdata['status'] = 1
    csvdata['msg'] = 'Csv file loaded okay'
    return csvdata

def convertToDic_renamed(columns, row, headersmapping):
    datadic = {}
    if len(columns)!=len(row):
        return datadic
    
    for index, item in enumerate(row):
        key = columns[index]
        newkey = headersmapping[key]
        datadic[newkey] = item
        
    return datadic
      
def convertToDicList_renamed(columns, rows, headersmapping):
    datadiclist = []
    for row in rows:
        datadic = convertToDic_renamed(columns, row, headersmapping)
        datadiclist.append(datadic)
    
    return datadiclist

def load_csvfile_diclist_renamed(csvfile, headersmapping):
    filename = csvfile.name
    csvdata = load_csvfile(csvfile)
    if csvdata['status']:
        logger.debug('csvfile loaded okay')
    else:
        svdata['diclist'] = []
        return csvdata
    
    columns = csvdata['columns']
    headersKnown = headersmapping.keys()
    if len(headersKnown)>0:
        n, missing = checkHeaders(headersKnown, columns)
        if n>0:
            msg = "Miss the following columns: " + ','.join(missing)
            csvdata['status'] = 0
            csvdata['msg'] = msg
            csvdata['diclist'] = []
            return csvdata
        
    rows = csvdata['rows']
    csv_diclist = convertToDicList_renamed(columns, rows, headersmapping)
    csvdata['diclist'] = csv_diclist
    csvdata['status'] = 1
    csvdata['msg'] = 'Csv file loaded okay'
    return csvdata

def load_excelfile_asbook(excelfile):
    workbook = openpyxl.load_workbook(excelfile, use_iterators = True, data_only=True)
    return workbook

def load_excelfile_1stSheet(excelfile, headersmapping):
    workbook = openpyxl.load_workbook(excelfile, read_only=True, data_only=True)
    worksheets = workbook.get_sheet_names()
    filedata = {}
    if len(worksheets)<1:
        filedata['status'] = 0
        filedata['msg'] = "No valid sheet is opened in the file: " + excelfile.name
        return filedata
    
    headersmappingLower = {}
    for key, value in headersmapping.items():
        keyLower = key.lower()
        headersmappingLower[keyLower] = value
    
    worksheet1 = worksheets[0]
    sheet = workbook.get_sheet_by_name(worksheet1)
    
    headersKnown = headersmapping.keys()
    diclist = []
    columns = []
    n = 0
    for row in sheet.iter_rows():
        n += 1
        if n==1:
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                columns.append(value)
            n0, missing = checkHeaders(headersKnown, columns)
            if n0>0:
                msg = "Miss the following columns: " + ','.join(missing)
                filedata['status'] = 0
                filedata['msg'] = msg
                filedata['diclist'] = []
                return filedata
        else:
            rowdic = {}
            i = 0
            for cell in row:
                key = columns[i]
                i += 1
                keyLower = key.lower()
                if keyLower in headersmappingLower:
                    newkey = headersmappingLower[keyLower]
                    value = cell.value
                    if value is None:
                        value = ''
                    rowdic[newkey] = value
                else:
                    logger.debug("key not found in the key mapping: ", key)
                
            diclist.append(rowdic)
            
    filedata['diclist'] = diclist
    filedata['status'] = 1
    msg = "From file: " + excelfile.name + " loaded " + str(n) + " rows"
    filedata['msg'] = msg
    return filedata

def load_excelfile_1stSheet_v2(excelfile):
    workbook = openpyxl.load_workbook(excelfile, read_only=True, data_only=True)
    worksheets = workbook.get_sheet_names()
    filedata = {}
    if len(worksheets)<1:
        filedata['status'] = 0
        filedata['msg'] = "No valid sheet is opened in the file: " + excelfile.name
        return filedata
    
    worksheet1 = worksheets[0]
    sheet = workbook.get_sheet_by_name(worksheet1)
    diclist = []
    columns = []
    n = 0
    for row in sheet.iter_rows():
        n += 1
        if n==1:
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                columns.append(value)
        else:
            rowdic = {}
            i = 0
            for cell in row:
                key = columns[i]
                i += 1
                if cell is not None:
                    value = cell.value
                    if value is None:
                        value = ''
                    rowdic[key] = value
            diclist.append(rowdic)
            
    filedata['diclist'] = diclist
    filedata['status'] = 1
    msg = "From file: " + excelfile.name + " loaded " + str(n) + " rows"
    filedata['msg'] = msg
    return filedata
    
def load_excelfile_1stSheet_anyformat(excelfile):
    workbook = openpyxl.load_workbook(excelfile, data_only=True)
    worksheets = workbook.sheetnames
    filedata = {}
    if len(worksheets)<1:
        filedata['status'] = 0
        filedata['msg'] = "No valid sheet is opened in the file: " + excelfile.name
        return filedata
    
    worksheet1 = worksheets[0]
    sheet = workbook[worksheet1]

    listlists = []
    n = 0
    for row in sheet.iter_rows():
        n += 1
        rowlist = []
        for cell in row:
            value = cell.internal_value
            rowlist.append(value)
        listlists.append(rowlist)
            
    filedata['listlists'] = listlists
    filedata['status'] = 1
    msg = "From file: " + excelfile.name + " loaded " + str(n) + " rows"
    filedata['msg'] = msg
    return filedata
    
def load_file(infile, headersmapping):
    filename = infile.name
    names = filename.split(".")
    suffix = names[1]
    if suffix.lower()=='csv':
        return load_csvfile_diclist_renamed(infile, headersmapping)
    else:
        return load_excelfile_1stSheet(infile, headersmapping)
    

def saveDiclistCSV(outcsvfile, diclist, fieldnames=[]):
    if len(diclist)==0:
        return
    
    if len(fieldnames)==0:
        fieldnames = diclist[0].keys()
        
    with open(outcsvfile, 'w') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(diclist) 
    
    
def load_excelfile(infile):
    filename = infile.name
    names = filename.split(".")
    suffix = names[1]
    return load_excelfile_1stSheet_v2(infile)
    
def modifyExcelDiclist(excelfile, headers, diclist, titleIn="sheet 1"):
    wb = openpyxl.load_workbook(filename=excelfile)
    if titleIn not in wb:
        return
    
    sheet = wb[titleIn]
    rowi = 0
    for index, item in enumerate(headers):
        ci = sheet.cell(row = (rowi+1), column = (index+1))
        ci.value = item
        
    for dici in diclist:
        rowi += 1
        for index, header in enumerate(headers):
            if header in dici:
                item = dici[header]
            else:
                item = ""
            ci = sheet.cell(row = (rowi+1), column = (index+1))
            ci.value = toString(item)

    wb.save(excelfile)
    
def modifyExcelCell(excelfile, rowNumber, colNumber, cellText, titleIn="sheet 1"):
    wb = openpyxl.load_workbook(filename=excelfile)
    if titleIn not in wb:
        return
    
    sheet = wb[titleIn]
    ci = sheet.cell(row = rowNumber, column = colNumber)
    ci.value = cellText
    wb.save(excelfile)
    
    
def saveExcelDiclist(excelfile, headers, diclist, titleIn="sheet 1", isNewSheet=True):
    if not isNewSheet:
        return modifyExcelDiclist(excelfile, headers, diclist, titleIn)
        
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.title = titleIn
    rowi = 0
    for index, item in enumerate(headers):
        ci = sheet.cell(row = (rowi+1), column = (index+1))
        ci.value = item
        
    for dici in diclist:
        rowi += 1
        for index, header in enumerate(headers):
            if header in dici:
                item = dici[header]
            else:
                item = ""
            ci = sheet.cell(row = (rowi+1), column = (index+1))
            ci.value = toString(item)

    wb.save(excelfile)
    
def reviseExcelDiclist(excelfile, headers, diclist, sheet_name="sheet 1"):
    wb = openpyxl.load_workbook(filename=excelfile)
    titleIn = sheet_name
    if titleIn not in wb:
        return
    
    idx = wb.sheetnames.index(sheet_name)
    ws = wb.get_sheet_by_name(sheet_name)
    wb.remove(ws)
    wb.create_sheet(sheet_name, idx)
    
    sheet = wb[titleIn]
    rowi = 0
    for index, item in enumerate(headers):
        ci = sheet.cell(row = (rowi+1), column = (index+1))
        ci.value = item
        
    for dici in diclist:
        rowi += 1
        for index, header in enumerate(headers):
            if header in dici:
                item = dici[header]
            else:
                item = ""
            ci = sheet.cell(row = (rowi+1), column = (index+1))
            ci.value = toString(item)

    wb.save(excelfile)
    
    
def loadSheet_all(sheet):
    diclist = []
    columns = []
    n = 0
    for row in sheet.iter_rows():
        empty = all(isinstance(cell, EmptyCell) for cell in row) 
        if empty:
            continue
        
        n += 1
        if n==1:
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                columns.append(value)
        else:
            rowdic = {}
            i = 0
            isEmpty = 0
            for cell in row:
                key = columns[i]
                i += 1
                if cell is not None:
                    value = cell.value
                    if value is None:
                        value = ''
                    else:
                        isEmpty = 1
                    rowdic[key] = value
                    
            if isEmpty==0:
                continue
            else:
                diclist.append(rowdic)
            
    sheetData = {'diclist':diclist, 'headers':columns}
    return sheetData
    
    
def loadSheet(sheet):
    sheetData = loadSheet_all(sheet)
    return sheetData['diclist']
    
def load_excelfile_asdic(excelfile):
    workbook = openpyxl.load_workbook(excelfile, read_only=True, data_only=True)
    worksheets = workbook.sheetnames
    
    filedata = {}
    if len(worksheets)<1:
        filedata['status'] = 0
        filedata['msg'] = "No valid sheet is opened in the file. "
        return filedata
    
    sheetnames = []
    for sheetname in worksheets:
        name = sheetname.upper().strip()
        sheetnames.append(name)
        sheet = workbook[sheetname]
        sheetData = loadSheet_all(sheet)
        filedata[name] = sheetData
    
    filedata['sheetnames'] = sheetnames
    if len(sheetnames)>0:
        filedata['status'] = 1
        filedata['msg'] = "Loaded file. "
    else:
        filedata['status'] = 0
        filedata['msg'] = "Error: No sheet loaded file. "
    return filedata
    
    
