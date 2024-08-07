#!/usr/bin/env python
import subprocess
import json
from subprocess import call

import string
import getpass
from urllib2 import urlopen, Request
import urllib2
import io
import argparse
import hashlib
import os, sys, re

SERVER = "http://test-server"
USERNAME = ''
PASSWORD = ''
USERPWD = " "
AUTH_URL = "/api/rest-auth/login/"

import openpyxl
from openpyxl.cell.read_only import EmptyCell
import datetime
import xlwt
import csv
import glob

SAMPLE_SHEET_NAMES = ["INSTRUCTIONS", "SAMPLES", "ASSAY", "ONTOLOGY"]
SAMPLE_SHEET_NAMES = ["INSTRUCTIONS", "SAMPLES", "ONTOLOGY"]
SAMPLE_ERRORCODE = {
    '101': 'Error: Excel file in incorrect format.',
    '102': 'Error: Assay sheet does not contain required sheet - ',
    '103': 'Error: Assay sheet does not contain valid "Instruction" sheet.',
    '104': 'Error: Assay sheet does not contain valid "Samples" sheet.',
    '105': 'Error: Sample type not identified in assay sheet - ',
    '106': 'Error: Excel file failed to load - ',
    '201': 'Error: Sample type not uniquely defined in database - ',
    '202': 'Error: Unknown attribute in assay sheet - ',
    '301': 'Error: Sample has an invalid Parent UID. ',
    '302': 'Error: Sample is missing data for either the "Name" or "File_PrimaryData" attribute - ',
    '303': 'Error: Sample has invalid entry for either the "Name" or "File_PrimaryData" attribute. ',
    '304': 'Error: Sample has no entry for the required "Scientist" attribute',
    '305': 'Error: "Scientist" name for the sample not registered in Seek: ',
    '401': 'Error: User has already uploaded a sample with this name to the database; please include the UID in order to update the sample metadata - ',
    '402': 'Error: Sample UID does not match sample name in the SEEK database - ',
    '403': 'Error: Sample name corresponds to more than one record in the database; please ask an admin for help.',
    '501': 'Error: No information provided for sample.',
    '502': 'Error: Required data is missing - ',
    '503': 'Error: Assay sheet is missing the "UID" attribute. ',
    '504': 'Error: Sample not saved into DB - ',
    '601': 'Warning: Sample not saved to the SEEK database - ',
    '602': 'Warning: Data file not associated with a sample in the SEEK database - '
}
DELIMITER_DBFIELD = "::"

def runQuery(url, server=None, token=None):
    if token is not None:
        return runAuthQuery(url, token, server)
    
    suffix = "' | grep -oPm250 '(?<=<d:href>)[^<]+'"
    serverIn = SERVER
    if server is not None:
        serverIn = server
    
    prefix = "curl " + USERPWD + " -s -X GET " + serverIn
    suffix = " -H 'accept: application/json'"
    apicmd = prefix + url + suffix
    resultset = subprocess.Popen([apicmd], stdout=subprocess.PIPE, shell=True
        ).communicate()[0].decode().split("\n")
    return resultset

def runAuthQuery(url, token, server=None):
    suffix = "' | grep -oPm250 '(?<=<d:href>)[^<]+'"
    serverIn = SERVER
    if server is not None:
        serverIn = server
        
    prefix = "curl -s -X GET " + serverIn
    suffix = " -H 'Authorization: Token " + token + "' -H 'accept: application/json'"
    apicmd = prefix + url + suffix
    resultset = subprocess.Popen([apicmd], stdout=subprocess.PIPE, shell=True
        ).communicate()[0].decode().split("\n")
    return resultset

def runAuthPost(url, jsondic, token, server=None):
    serverIn = SERVER
    if server is not None:
        serverIn = server
    
    prefix = "curl -i -X POST " + serverIn
    suffix = " -H 'Authorization: Token " + token + "'"
    suffix += " -H 'accept: application/json'"
    suffix += " -H 'Content-Type:application/json'"
    data = " -d '" + jsondic + "'"
    apicmd = prefix + url + suffix + data
    resultset = subprocess.Popen([apicmd], stdout=subprocess.PIPE, shell=True
        ).communicate()[0].decode().split("\n")
    
    resultdDic = json.loads(resultset[-1])
    status = resultdDic['status']
    msg = resultdDic['msg']
    data = resultdDic['data']
    if status:
        print("Sample submission successful")
    else:
        print("Sample submission with error: " + msg)
    
    return resultdDic

def getDseqFiles(sample_id):
    try:
        id = int(sample_id)
        url = "/api/samples/" + str(id) + "/"
        search = 0
    except:
        uid = sample_id
        url = "/api/samples/?search=" + str(uid)
        search = 1
    
    resultset = list(runQuery(url))
    dic0 = json.loads(resultset[0])
    if search==1:
        dic0 = dic0[0]
    for key, value in dic0.items():
        if "file_primarydata" in key:
            filename = value
            file_uid, file_link = searchDatafileUID(filename)
            if len(file_link)>0:
                downloadFile(file_link)
                
    filef = ""
    if "file_primarydata_forward" in dic0:
        filef = dic0["file_primarydata_forward"]
    
    filer = ""
    if "file_primarydata_forward" in dic0:
        filer = dic0["file_primarydata_reverse"]
        
    return filef, filer

def parseDatafile(df_dic, ifDownload=False):
    df_uid = ''
    if not isinstance(df_dic, dict):
        print('Error: No data file info is retrieved from DB.')
        return df_uid
        
    if 'status' not in df_dic:
        print('Error: Data file info incomplete.')
        return df_uid
        
    status = df_dic['status']
    if not status:
        print('Error: Data file info incorrect: %s'%df_dic['msg'])
        return df_uid
        
    df_uid = df_dic['uid']
    print("Data file has valid UID: %s "%df_uid)
    return df_uid

def parseExternalDatafile(filename, filelink):
    msg = 'To be implemented'
    status = 0
    if PRIDE_URL in filelink:
        from api_calls_pride import verifyPrideFileDownload
        msg, status = verifyPrideFileDownload(filename)
        
    return msg, status

def searchSampleFiles(sample_keyword, server, token):
    fileDicList = []
    
    if len(str(sample_keyword))==0:
        print("No sample found: empty keyword")
        return fileDicList
    
    url = "/api/samples/?search=" + str(sample_keyword)
    resultset = list(runQuery(url, server, token))
    if len(resultset)==0:
        print("No sample found: no match")
        return fileDicList
    
    jsonstr = resultset[0]
    dicList = json.loads(jsonstr)
    if len(dicList)==0:
        print("No sample found: no match")
        return fileDicList
    
    for dici in dicList:
        fileDic = {}
        for key, value in dici.items():
            if "file_" in key:
                df_dic = value
                if isinstance(df_dic, dict):
                    df_uid = parseDatafile(df_dic)
                else:
                    filelink_key = key.replace('file_','link_')
                    if filelink_key in dici:
                        filename = value
                        filelink = dici[filelink_key]
                        print("Data file is external: %s "%filelink) 
                        msg, status = parseExternalDatafile(filename, filelink)
                    else:
                        print("Data file is not external: %s "%filelink)
                
                fileDic[key] = df_uid
        fileDic['sample_uid'] =  dici['uid']
        fileDicList.append(fileDic)
        
    print("Number of samples found: %d"%len(fileDicList))
    return fileDicList
    
def querySampleFiles(sample_id, server, token):
    fileDic = {}
    try:
        id = int(sample_id)
        url = "/api/samples/" + str(id) + "/"
    except:
        print("No sample found: not valid sample id")
        return fileDic
    
    resultset = list(runQuery(url, server, token))
    if len(resultset)==0:
        print("No sample found: no match")
        return fileDic
    
    jsonstr = resultset[0]
    sample_dic = json.loads(jsonstr)
    if not sample_dic:
        print("No sample found: empty sample meta-data")
        return fileDic
    
    if 'sample_uid' not in sample_dic:
        print("No sample found: sample uid not available")
        return fileDic

    for key, value in sample_dic.items():
        if "file_" in key:
            fileDic[key] = value

    fileDic['sample_uid'] =  sample_dic['uid']
    return fileDic    
    
    
def getSampleFiles(sample_id, server=None, token=None):
    try:
        id = int(sample_id)
        url = "/api/samples/" + str(id) + "/"
        fileDic = querySampleFiles(id, server, token)
        fileDicList = [fileDic]
    except:
        sample_keyword = sample_id
        url = "/api/samples/?search=" + str(sample_keyword)
        fileDicList = searchSampleFiles(sample_keyword, server, token)

    return fileDicList

def searchDatafileUID(datafilename):
    if len(datafilename.strip())==0:
        return "", ""
    
    url = "/api/datafiles/?search=" + str(datafilename.strip())
    resultset = runQuery(url)
    dic0 = json.loads(resultset[0])
    dic0 = dic0[0]
    datafileuid = ""
    datafilelink = ""
    if "title" in dic0:
        datafileuid = dic0["title"]
        datafilelink = dic0['link']
        
    print("datafileuid: ", datafileuid)
    print("datafilelink: ", datafilelink)
    return datafileuid, datafilelink

def downloadFile(url):
    file_name = url.split('/')[-1]
    u = urllib2.urlopen(url)
    f = open(file_name, 'wb')
    meta = u.info()
    file_size = int(meta.getheaders("Content-Length")[0])
    print("Downloading: %s Bytes: %s" % (file_name, file_size))

    file_size_dl = 0
    block_sz = 8192
    while True:
        buffer = u.read(block_sz)
        if not buffer:
            break

        file_size_dl += len(buffer)
        f.write(buffer)
        status = r"%10d  [%3.2f%%]" % (file_size_dl, file_size_dl * 100. / file_size)
        status = status + chr(8)*(len(status)+1)

    f.close()
    
def dwonloadDF(url, filename, username, password):
    import urllib2
    req = urllib2.Request(url)
    base64string = base64.encodestring('%s:%s' % (username, password))[:-1]
    authheader =  "Basic %s" % base64string
    req.add_header("Authorization", authheader)
    res = urllib2.urlopen(req)
    with open(filename,'wb') as output:
        output.write(res.read())

def submitSalmonJob(fileDownload):
    filef = ""
    if "file_primarydata_forward" in fileDownload:
        filef = fileDownload["file_primarydata_forward"]
    
    filer = ""
    if "file_primarydata_forward" in fileDownload:
        filer = fileDownload["file_primarydata_reverse"]
        
    if len(filef)==0 or len(filer)==0:
        print("Exit: not paired reads")
        return

    if "_1_sequence.fastq" not in filef:
        print("Exit: fastq file for forward reads not in standard format: ", filef)
        return
    
    fileprefixf = filef.replace("_1_sequence.fastq", "")
    
    if "_2_sequence.fastq" not in filer:
        print("Exit: fastq file for reverse reads not in standard format: ", filer)
        return
    
    fileprefixr = filer.replace("_2_sequence.fastq", "")
    
    if fileprefixf!=fileprefixr:
        print("Exit: fastq files for forward and reverse reads not match")
        return
    
    prefix = "sbatch ./salmon.sh "
    apicmd = prefix + fileprefixr
    try:
        call([apicmd], shell=True)
        status = True
        print(apicmd)
    except:
        status = False
        print("Failed in submitting Salmon alignment: " + apicmd)
    return status

def getAuthToken(username, password, server=None):
    if username is None:
        print("Error: User name is not valid")
        return None
    if password is None:
        print("Error: password is not valid")
        return None
    
    serverIn = SERVER + AUTH_URL
    if server is not None:
        serverIn = server + AUTH_URL
    
    prefix = "curl -X POST -d "
    pwd = '"username": "' + username + '","password": "' + password + '"'
    pwd = "'{" + pwd + "}'"
    suffix = " -H 'Content-Type: application/json' "
    apicmd = prefix + pwd + suffix + serverIn
    resultset = subprocess.Popen([apicmd], stdout=subprocess.PIPE, shell=True
        ).communicate()[0].decode().split("\n")
    
    dic0 = json.loads(resultset[0])
    token = ''
    if 'key' in dic0:
        token = dic0['key']
    elif 'auth_token' in dic0:
        token = dic0['token']
        
    print("Auth token: %s"%token)
    return token

def runSalmonWorkflow(sampleid, server):
    fileDicList = getSampleFiles(sampleid, server)
    
    for fileDic in fileDicList:
        if 'sample_uid' in fileDic:
            print('sample_uid: %s'% fileDic['sample_uid'])
        else:
            continue
        
        fileDownload = {}
        n = 0
        for key, value in fileDic.items():
            if "file_" in key:
                datafile_uid = value
                if datafile_uid in fileDic:
                    datafile_link = fileDic[datafile_uid]
                    if len(datafile_link)>0:
                        downloadFile(datafile_link)
                        fileDownload[key] = datafile_uid
        submitSalmonJob(fileDownload)
        
def runAuthSampleSubmission(sampledata, sampleType, username, password, server=None):
    token = getAuthToken(username, password, server)
    url = "/api/sampleupload/"
    sdata = json.loads(sampledata)
    
    uid = None
    if 'UID' in sdata:
        uid = sdata['UID']
    elif 'uid' in sdata:
        uid = sdata['uid']
        
    if uid is not None:
        if len(uid.strip())>0:
            print("Update operation")
            jsondic = sampledata
            resultdic = runAuthPost(url, jsondic, token)
            return
        
    data = {}
    data['User'] = {'username':username, 'user_id':2, 'projectid':3, 'lababbv':'BIO'}
    data['Samples'] = sdata
    data['Sample type'] = sampleType
    jsondic = json.dumps(data)
    resultdic = runAuthPost(url, jsondic, token)
    return

def loadSheet_all(sheet):
    diclist = []
    columns = []
    n = 0
    for row in sheet.iter_rows():
        empty = all(isinstance(cell, EmptyCell) for cell in row) # check if all cells are empty
        if empty:
            continue
        
        n += 1
        if n==1:
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                columns.append(value)
        else:
            rowdic = {}
            i = 0
            isEmpty = 0
            for cell in row:
                key = columns[i]
                i += 1
                if cell is not None:
                    value = cell.value
                    if value is None:
                        value = ''
                    else:
                        isEmpty = 1
                    rowdic[key] = value
                    
            if isEmpty==0:
                continue
            else:
                diclist.append(rowdic)
            
    sheetData = {'diclist':diclist, 'headers':columns}
    return sheetData
 
def load_excelfile_asdic(excelfile):
    workbook = openpyxl.load_workbook(excelfile, read_only=True, data_only=True)
    worksheets = workbook.sheetnames
    
    filedata = {}
    if len(worksheets)<1:
        filedata['status'] = 0
        filedata['msg'] = "No valid sheet is opened in the file. "
        return filedata
    
    sheetnames = []
    for sheetname in worksheets:
        name = sheetname.upper().strip()
        sheetnames.append(name)
        sheet = workbook[sheetname]
        sheetData = loadSheet_all(sheet)
        filedata[name] = sheetData
        
    print("loaded excelfile asdic: %s"%excelfile)
    filedata['sheetnames'] = sheetnames
    if len(sheetnames)>0:
        filedata['status'] = 1
        filedata['msg'] = "Loaded file. "
    else:
        filedata['status'] = 0
        filedata['msg'] = "Error: No sheet loaded file. "
        print(filedata['msg'])
    return filedata

def loadSampleTypes(diclist_instruction):
        sampleTypes = {}      
        sampleTypes_order = []
        sampleHeaders = []
        for dici in diclist_instruction:
            if "Field" not in dici or "Database Field" not in dici:
                msg = "Error: Instruction sheet should contain 'Field' or 'Database Field' columns."
                print(msg)
                return {}, []
                
            tbheader = dici["Field"]
            if tbheader is not None:
                header = str(tbheader)
                if len(header.strip())==0:
                    continue
            else:
                continue
            
            dbfield = dici["Database Field"]
            if DELIMITER_DBFIELD not in dbfield:
                msg = "Error: 'Database Field' column should follow the format 'tableName::attributeName."
                print(msg)
                return {}, []
                
            terms = dbfield.split(DELIMITER_DBFIELD)     
            sampleType = terms[0]         
            attribute = terms[1]            
            if sampleType in sampleTypes:
                attributeMapping = sampleTypes[sampleType]
            else:
                attributeMapping = {}
                                
            attributeMapping[tbheader] = attribute
            sampleTypes[sampleType] = attributeMapping
            sampleHeaders.append(tbheader)
            if sampleType not in sampleTypes_order:
                sampleTypes_order.append(sampleType)
            
        return sampleTypes, sampleTypes_order, sampleHeaders

def splitSampleTypes(sampleTypes, diclist_samples):
        sample_sheets = {}
        for sampleType in sampleTypes:
            sample_sheets[sampleType] = []
            
        unique_samples = {}
        for dici_meta in diclist_samples:
            for sampleType, attributeMapping in sampleTypes.items():
                diclist = sample_sheets[sampleType]
                dici_sample = {}
                samplename = None
                for header, value in dici_meta.items():
                    if header in attributeMapping:
                        attribute = attributeMapping[header]
                        dici_sample[attribute] =  value
                        
                        if attribute.lower()=='name':
                            samplename = value
                        elif attribute.lower()=='file_primarydata':
                            samplename = value
                        elif attribute.lower()=='file_primarydata_forward':
                            samplename = value
                        elif attribute.lower()=='file_primarydata_reverse':
                            samplename = value
                            
                if samplename is None:
                    msg = 'Error: one sample does not have an unique identifier, which will be captured later on in __getRecord()'
                    print(msg)
                elif samplename in unique_samples:
                    dici_sample = unique_samples[samplename]
                else:
                    unique_samples[samplename] = dici_sample
                        
                diclist.append(dici_sample)
                sample_sheets[sampleType] = diclist
        return sample_sheets

def batchUpload(infile, feedbackfile, user_seek=None):
        msg = "batchUpload"
        status = 0
        
        try:
            filedata = load_excelfile_asdic(infile)
        except:
            msg = SAMPLE_ERRORCODE['101']
            status = 0
            print(msg)
            return msg, status, None
        
        status = filedata['status']
        msg = filedata['msg']
        if status==0:
            msg = SAMPLE_ERRORCODE['106'] + msg
            print(msg)
            return msg, status, None
        
        status = 1
        for sheetname in SAMPLE_SHEET_NAMES:
            msg = 'Error: Assay sheet does not contain required sheet - '
            if sheetname not in filedata['sheetnames'] or sheetname not in filedata:
                status = 0
                msg += sheetname + ';'
            
            if status==0:
                print(msg)
                return msg, status, None
        
        sheetData_ins = filedata["INSTRUCTIONS"]
        diclist_ins = sheetData_ins['diclist']
        sampleTypes, sampleTypes_order, sampleHeaders = loadSampleTypes(diclist_ins)
        if len(sampleTypes.keys())==0:
            msg = SAMPLE_ERRORCODE['103']
            status = 0
            print(msg)
            return msg, status, None
        
        return sampleHeaders, 1, diclist_ins
    

def outputUploadSamples_V2(diclist, diclist_feedback, headers, feedbackfile):
        book = xlwt.Workbook(encoding="utf-8")
        sheet1 = book.add_sheet("Samples")
        
        row = 0
        for index, header in enumerate(headers):
            try:
                newitem = toString(header)
            except:
                newitem = cleanString(header)
            sheet1.write(row, index, newitem)
        
        style = xlwt.easyxf('pattern: pattern solid, fore_colour red;')
        style0 = xlwt.Style.easyxf('pattern: pattern solid, fore_colour white;')
 
        i = 0
        for dici in diclist:
            dici_feedback = diclist_feedback[i]
            row += 1
            for index, header in enumerate(headers):
                if header in dici:
                    newitem = dici[header]
                else:
                    newitem = ""
                
                try:
                    newitem = str(newitem)
                except:
                    newitem = toString(newitem)
                
                if header in dici_feedback:
                    feedback = dici_feedback[header]
                    if feedback is None:
                        sheet1.write(row, index, newitem)
                    elif 'Error' in toString(feedback):
                        feedback = feedback + ":" + newitem
                        sheet1.write(row, index, feedback, style)
                    elif 'UID' in header:
                        sheet1.write(row, index, feedback)
                    else:
                        sheet1.write(row, index, newitem)
                else:
                    sheet1.write(row, index, newitem)
                
            i += 1 
        book.save(feedbackfile)
        
def toString(itemIn):
    if itemIn is None:
        itemOut = ' '
        return itemOut
    
    try:
        itemOut = str(itemIn)
    except:
        strtype = type(itemIn)
        if strtype == unicode:        
            itemOut = itemIn.encode("utf-8")
            itemOut = str(itemOut)
        else:
            itemOut = itemIn
            
    return itemOut.strip()        
        
def cleanString(itemIn):
    if type(itemIn)==unicode:
        itemOut = itemIn.encode("utf-8")
        return itemOut
        
    newitem = unicode(itemIn, "utf-8", errors="ignore")
    itemOut = newitem
    return itemOut
        
def outputUploadSamples(jsondata, headers, outputfile):
        if "samples" not in jsondata:
            msg = "Error: no valid samples.json file for parsing"
            status = 0
            print(msg)
            return msg, status
        
        samples = jsondata["samples"]
        book = xlwt.Workbook(encoding="utf-8")
        sheet1 = book.add_sheet("Samples")
        
        row = 0
        for index, header in enumerate(headers):
            try:
                newitem = toString(header)
            except:
                newitem = cleanString(header)
            sheet1.write(row, index, newitem)
        
        style = xlwt.easyxf('pattern: pattern solid, fore_colour red;')
        style0 = xlwt.Style.easyxf('pattern: pattern solid, fore_colour white;')
        i = 0
        for sample in samples:
            row += 1
            for index, header in enumerate(headers):
                if header=='Sample ID':
                    newitem = sample['sample_id']
                    print('Sample ID: %s'%newitem)
                elif header in jsondata:
                    newitem = jsondata[header]
                elif header in sample:
                    newitem = sample[header]
                else:
                    newitem = ""
                
                try:
                    newitem = str(newitem)
                except:
                    newitem = toString(newitem)
                
                sheet1.write(row, index, newitem)
                
            i += 1 
        book.save(outputfile)        

def outputDiclist(diclist, headers, outputfile, sheetname):
    book = xlwt.Workbook(encoding="utf-8")
    sheet1 = book.add_sheet(sheetname)
        
    row = 0
    for index, header in enumerate(headers):
        try:
            newitem = toString(header)
        except:
            newitem = cleanString(header)
        sheet1.write(row, index, newitem)
        
    style = xlwt.easyxf('pattern: pattern solid, fore_colour red;')
    style0 = xlwt.Style.easyxf('pattern: pattern solid, fore_colour white;')
    i = 0
    for dici in diclist:
        row += 1
        for index, header in enumerate(headers):
            if header in dici:
                newitem = dici[header]
            else:
                newitem = ""
                
            try:
                newitem = str(newitem)
            except:
                newitem = toString(newitem)
                
            sheet1.write(row, index, newitem)
                
        i += 1 
    book.save(outputfile)
    print("Number of rows output: %d"%i)
    
def parseInfosite(rstfile):
    print("Load infosite from: %s"%rstfile)
    lines = open(rstfile).readlines()
    file = rstfile
    sample_data = []
    for line in lines:
        line = line.strip()
    
        if not line.startswith("|"):
            continue

        if len(line.replace("|", "").replace(" ", "")) == 0:
            continue

        if len(line.split("|")[1].strip()) == 0:
            continue

        numberItems = len(line.split("|"))

        if numberItems == 14:
            sample_data.append(line)        

    count = 0
    diclist1 = []
    headers1 = []
    row = None
    for sample in sample_data:
        row = []
        line = sample[1:-1].split('|')
        if count == 0:
            for entry in line:
                entry = re.sub(r'\(% of lane\)', ',Read % of lane', entry)
                entry = re.sub(r'\(%\)', ',Aligned % of reads', entry)
                entry = re.sub(r'Contam', 'Contam,ContamPercent', entry)
                header = entry.strip()[1:-1]
                if "," in header:
                    terms = header.split(",")
                    for term in terms:
                        row.append(term.strip())
                        headers1.append(term.strip())
                else:
                    row.append(header)
                    headers1.append(header)
    
            count += 1
        else:
            if "Project ID" in sample and "Sample ID" in sample:
                continue

            i = 0
            dici = {}
            for entry in line:
                entry = re.sub(r'\(', ' - ', entry)
                entry = re.sub(r'\)', '', entry)
                entry = re.sub(r'_|`|,|%', '', entry)
                entry = re.sub(r':boldred:', '', entry)
                entry = re.sub(r' - ', ',', entry)
                value = entry.strip()
                if "," in value:
                    terms = value.split(",")
                    for term in terms:
                        row.append(term.strip())
                        header = headers1[i]
                        dici[header] = term.strip()
                        i += 1
                else:
                    row.append(value)
                    header = headers1[i]
                    dici[header] = value
                    i += 1
            diclist1.append(dici) 

    sample_data = []
    for line in lines:
        line = line.strip()
        
        if not line.startswith("|"):
            continue

        if len(line.replace("|", "").replace(" ", "")) == 0:
            continue

        if len(line.split("|")[1].strip()) == 0:
            continue

        numberItems = len(line.split("|"))

        if numberItems == 21:
            sample_data.append(line)        

    count = 0
    diclist2 = []
    headers2 = []
    row = None
    for sample in sample_data:
        row = []
        line = sample[1:-1].split('|')
        if count == 0:
            for entry in line:
                entry = re.sub(r'\([^()]*\)', '', entry)
                entry = re.sub(r'Contam', 'Contam,ContamPercent', entry)
                header = entry.strip()[1:-1]
                
                if "," in header:
                    terms = header.split(",")
                    for term in terms:
                        row.append(term.strip())
                        headers2.append(term.strip())
                else:
                    row.append(header)
                    headers2.append(header)

            count += 1
 
        else:
            if "Project ID" in sample and "Sample ID" in sample:
                continue
            
            i = 0
            dici = {}
            for entry in line:
                entry = re.sub(r'\([^()]*\)', '', entry)
                entry = re.sub(r'_|`|,|%', '', entry)
                entry = re.sub(r' - ', ',', entry)
                
                value = entry.strip()
                if "," in value:
                    terms = value.split(",")
                    for term in terms:
                        row.append(term.strip())
                        header = headers2[i]
                        dici[header] = term.strip()
                        i += 1
                else:
                    row.append(value)
                    header = headers2[i]
                    dici[header] = value
                    i += 1
            diclist2.append(dici) 
    data = {'diclist1':diclist1, 'headers1':headers1, 'diclist2':diclist2, 'headers2':headers2}
    return data

def startJVM():
    import jaydebeapi
    import jpype
    jar = '/net/bmc-pub15/data/bmc/Code/pipeline_1.0.0/tools/fm_integration/lib/fmjdbc.jar'
    args='-Djava.class.path=%s' % jar
    jvm_path = jpype.getDefaultJVMPath()
    jpype.startJVM(jvm_path, args)
    URL = 'jdbc:filemaker://bmc-5.mit.edu/DNA_db'
    conn = jaydebeapi.connect('com.filemaker.jdbc.Driver',URL, ['bmc','password'], jar,)
    curs = conn.cursor()
    return conn, curs

def retrieveFMDBSamples(curs, tableName, fieldNameIn, fieldValueIn):
    query = "select * from FileMaker_Fields where TableName='" + tableName + "'"
    curs.execute(query)
    rows = curs.fetchall()
    fieldNames = []
    for row in rows:
        fieldName = row[1]
        fieldNames.append(fieldName)
    
    query = "select * from " + tableName + " where " + fieldNameIn + "='" + fieldValueIn + "'"
    curs.execute(query)
    diclist = []
    rows = curs.fetchall()
    for row in rows:
        dici = {}
        for i, fieldName in enumerate(fieldNames):
            dici[fieldName] = row[i]
        
        diclist.append(dici)
        
    return diclist

def loadTemplate(templetfile):
    print("templetfile: %s"%templetfile)
    sampleHeaders, status, sampleMapping = batchUpload(templetfile, None, None)
    return sampleHeaders, status, sampleMapping

def loadJson(jsonfile):
    print("Loading jsonfile: %s"%jsonfile)
    jsondata = {}
    with open(jsonfile, 'r') as f:
        jsondata = json.load(f)
    
    if "samples" not in jsondata:
        msg = "Error: not a valid samples.json file: " + jsonfile
        status = 0
        print(msg)
        return {}

    return jsondata

def getFileChecksum(fullfilename, checksumFormat='MD5'):
    cf = checksumFormat.upper()
    if cf=='MD5':
        fi = open(fullfilename,'rb').read()
        print('open file: %s'%fullfilename)
        print'Calculate MD5 checksum...'
        md5 = hashlib.md5(fi).hexdigest()
        checksum = md5
        print('MD5 checksum: %s'%md5)
    elif cf=='SHA1':
        openedFile = open(fullfilename,'rb')
        readFile = openedFile.read()
        sha1Hash = hashlib.sha1(readFile)
        sha1Hashed = sha1Hash.hexdigest()
        checksum = sha1Hashed
    else:
        checksum = 'NA'
    return checksum

def getFileInfo(flowcell_path, sample):
    terms = flowcell_path.split('_')
    flowcellID = terms[-1]  
    sample_path = sample['publish_path'] + '/' + sample['project_id'] + '/' + sample['sample_id'] + '-' + flowcellID
    prefix = sample['project_id'] + '_' + sample['sample_id'] + '_NA_sequence'
    
    singleEndFile = sample_path + '/' + prefix + '.*'
    fileList = glob.glob(singleEndFile)
    if len(fileList)==1:
        fileInfo = {'sequencingtype':'single'}
        fullfilename = fileList[0]
        checksum = getFileChecksum(fullfilename, 'MD5')
        terms = fullfilename.split("/")
        fastqfile = terms[-1]
        terms = fastqfile.split(".")
        filetype = terms[-1]
        dici = {'filename':fastqfile, 'link':fullfilename, 'filetype':filetype, 'checksum':checksum}
        diclist = [dici]
        fileInfo['filelist'] = diclist
        return fileInfo            
    
    pairedEndFile = sample_path + '/' + sample['project_id'] + '_' + sample['sample_id'] + '_NA_sequence_*.*'
    fileList = glob.glob(pairedEndFile)
    if len(fileList)==2:
        fileInfo = {'sequencingtype':'paired'}
        diclist = []
        for fullfilename in fileList:
            checksum = getFileChecksum(fullfilename, 'MD5')
            terms = fullfilename.split("/")
            fastqfile = terms[-1]
            terms = fastqfile.split(".")
            filetype = terms[-1]
            dici = {'filename':fastqfile, 'link':fullfilename, 'filetype':filetype, 'checksum':checksum}
            diclist.append(dici)
        
        fileInfo['filelist'] = diclist
        return fileInfo 
    
    fileInfo = {'sequencingtype':'undefined'}
    fileInfo['filelist'] = []
    return fileInfo 

def getInfositeRST(flowcell_path, sample):
    terms = flowcell_path.split('_')
    flowcellID = terms[-1]      
    terms = flowcell_path.split('/')
    flowcellName = terms[-1]
    infosite_path = sample['publish_path'] + '/' + sample['project_id'] + '/' + 'infosite-' + flowcellID
    rstfile = flowcellName + '_infosite.rst'
    fullrstfile = infosite_path + '/' + rstfile
    return infosite_path, fullrstfile

def getSamples(jsondata, headers, projectid):
    samplesDiclist = []
    if "samples" not in jsondata:
        msg = "Error: no valid samples.json file for parsing"
        status = 0
        print(msg)
        return msg, status, samplesDiclist
        
    samples = jsondata["samples"]
    flowcell_path = jsondata["flowcell_path"]
 
    i = 0
    for sample in samples:
        project_id = sample["project_id"]
        if projectid is not None and projectid!=project_id:
            continue
                    
        dici = {}      
        for index, header in enumerate(headers):
            if header in jsondata:
                newitem = jsondata[header]
            elif header in sample:
                newitem = sample[header]  
            elif header=='Sample ID':
                newitem = sample['sample_id']
            else:
                newitem = ""
            try:
                newitem = str(newitem)
            except:
                newitem = toString(newitem)
                
            dici[header] = newitem
        
        fileInfo = getFileInfo(flowcell_path, sample)
        if fileInfo['sequencingtype']=='single':
            info = fileInfo['filelist'][0]
            dici['Forward Read File Name'] = info['filename']
            dici['Forward Read File Link'] = info['link']
            dici['Forward File Checksum'] = info['checksum']
            dici['sequencing file type'] = info['filetype']
        elif fileInfo['sequencingtype']=='paired':
            info = fileInfo['filelist'][0]
            dici['Forward Read File Name'] = info['filename']
            dici['Forward Read File Link'] = info['link']
            dici['Forward File Checksum'] = info['checksum']
            dici['sequencing file type'] = info['filetype']
            
            info = fileInfo['filelist'][1]
            dici['Reverse Read File Name'] = info['filename']
            dici['Reverse Read File Link'] = info['link']
            dici['Reverse File Checksum'] = info['checksum']
            dici['sequencing file type'] = info['filetype']
                
        samplesDiclist.append(dici)    
        i += 1 
    
    status = 1
    msg = 'Number of sample loaded: ' + str(i)
    print(msg)
    return msg, status, samplesDiclist
 
def csv_dict_list(incsvfile):
    reader = csv.DictReader(open(incsvfile, 'rb'))
    dict_list = []
    for line in reader:
        dict_list.append(line)
    return dict_list

def loadInfosite(samplesDiclist, infositecsvfile, identifier):
    infoDiclist = csv_dict_list(infositecsvfile)
    if identifier is None:
        identifier = 'sample_id'
        
    infodic = diclistToDicdic(infoDiclist, identifier, '') 
    diclist = appendSampleInfo(samplesDiclist, identifier, infodic)
    return diclist

def appendSampleInfo(samplesDiclist, identifier, samplesDic):
    diclist = []
    i = 0
    for dici in samplesDiclist:
        if identifier not in dici:
            diclist.append(dici)
            continue
        
        id = dici[identifier]
        if id not in samplesDic:
            for key in samplesDic.keys():
                if key in id:
                    id = key
                    
        if id not in samplesDic:    
            diclist.append(dici)
            print("Identifier not found: %s"%id)
            continue
            
        sampleDic = samplesDic[id]
        for key, value in sampleDic.items():
            if key in dici:
                dici[key] = value
        diclist.append(dici)
        i += 1
        
    print("Number of samples appened: %d"%i)
    return diclist


def diclistToDicdic(diclist, identifier, keyprefix='', keyRemove=None):
    dicDic = {}
    i = 0
    for dici in diclist:
        if identifier not in dici:
            continue
        
        id = dici[identifier]
        if keyRemove is not None:
            id = id.replace(keyRemove, '')
        
        dici_new = {}
        for key, value in dici.items():
            newKey = keyprefix + key
            dici_new[newKey] = value
        
        dicDic[id] = dici_new
        i += 1
        
    print("Number of FMDB records converted: %s"%i)
    return dicDic
    
def getFMDBSamples(curs, samplesDiclist, tableName, fieldName, fieldValue):
    samplesFMDB = retrieveFMDBSamples(curs, tableName, fieldName, fieldValue)
    identifier = 'SampleName'
    keyprefix = tableName + "::"
    samplesDic = diclistToDicdic(samplesFMDB, identifier, keyprefix)
    return samplesDic

 
def loadInfositeFromRST(jsondata, project_id, samplesDiclist):
    flowcell_path = jsondata["flowcell_path"]
    samples = jsondata["samples"]
    sample0 = samples[0]
    infosite_path, fullrstfile = getInfositeRST(flowcell_path, sample0)
    data = parseInfosite(fullrstfile)
    terms = flowcell_path.split('_')
    flowcellID = terms[-1]      # such as "5337G"
    
    # first section of infosite
    diclist1 = data['diclist1']
    print("Number of samples loaded from infosite: %d"%len(diclist1))
    identifier = 'Sample ID'
    keyprefix = ''
    keyRemove = "-" + flowcellID
    
    samplesDic = diclistToDicdic(diclist1, identifier, keyprefix, keyRemove)
    identifier = "sample_id"
    samplesDiclist = appendSampleInfo(samplesDiclist, identifier, samplesDic)
    
    # second section of infosite
    diclist2 = data['diclist2']
    print("Number of samples loaded from infosite: %d"%len(diclist2))
    identifier = 'Sample ID'
    
    keyprefix = ''
    keyRemove = "-" + flowcellID
    samplesDic = diclistToDicdic(diclist2, identifier, keyprefix, keyRemove)
    identifier = "sample_id"
    samplesDiclist = appendSampleInfo(samplesDiclist, identifier, samplesDic)
    
    #print(samplesDic.keys())
    return samplesDiclist

def adjustSamples(samplesDiclist):
    for sample in samplesDiclist:
        typeNotes = sample["RNA_samples::SampleNotes"]
        if ";" in typeNotes:
            terms = typeNotes.split(";")
            sample["RNA_samples::SampleType"] = terms[0].replace('Type=', '')
            sample["RNA_samples::SampleNotes"] = terms[1].replace('Notes=', '')
            
        for key, value in sample.items():
            if "::" in key:
                terms = key.split("::")
                field = terms[1]
                if field in sample:
                    sample[key] = sample[field]
    
        id = sample["RNA_samples::RNA_ID"]
        sample["RNA_samples::RNA_ID"] = str(int(id))
    return samplesDiclist

def parseSamples(templetfile, jsonfile, projectid, outputfile, infositecsvfile, useAPI=True, submitDfile=True):
    jsondata = loadJson(jsonfile)
    
    if outputfile is None:
        names = templetfile.split(".")
        n = len(names)
        datenow = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
        outputfile = '.'.join(names[:(n-1)]) + '_samples-' + datenow + '.xls'
        print("outputfile: %s"%outputfile)
    
    sampleHeaders, status, diclist_ins = batchUpload(templetfile, None, None)
    if status==0:
        print(sampleHeaders)
        print("Error: no putput file is generated")
        return
    
    msg, status, samplesDiclist = getSamples(jsondata, sampleHeaders, projectid)
    if not status:
        print(msg)
        return
    
    conn, curs = startJVM()
    samplesDic = getFMDBSamples(curs, samplesDiclist, "RNA_samples", "ProjectID", projectid)
    identifier = "sample_name"
    samplesDiclist = appendSampleInfo(samplesDiclist, identifier, samplesDic)
    if infositecsvfile is not None:
        identifier = 'sample_id'
        samplesDiclist = loadInfosite(samplesDiclist, infositecsvfile, identifier)
    else:
        samplesDiclist = loadInfositeFromRST(jsondata, projectid, samplesDiclist)       
        
    samplesDiclist = adjustSamples(samplesDiclist)
    if useAPI:
        from api_sampleSubmit import runAuthSampleBatchSubmission, getAuthToken
        username = 'user'
        password = 'password'
        server = "http://test-server"
        
        userid = 3
        projectid = 1
        lababbv = 'BMC'
        token = getAuthToken(username, password, server)
        userInfo = {'username':username, 'user_id':userid, 'projectid':projectid, 'lababbv':lababbv}
        samplesDiclist = runAuthSampleBatchSubmission(samplesDiclist, diclist_ins, token, userInfo, submitDfile)
    
    
    outputDiclist(samplesDiclist, sampleHeaders, outputfile, "Samples")
 
        
def main():
    parser = argparse.ArgumentParser(description="Parse samples.json and infosite.csv to fill the assaysheet sheet",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    
    parser.add_argument('jsonfile', help='Input file in json format for a list of samples retrieved from FileMaker')
    
    parser.add_argument('-t', '--templetfile', 
                        help='Mandatary, the template assaysheet in Excel format.')
        
    parser.add_argument('-p', '--projectid', 
                        help='Optional, only output those samples in the project if projectid is valid.', default=None)        

    parser.add_argument('-i', '--infositecsvfile', 
                        help='Optional, a csv file with the info from the infosite.', default=None)        
        
    parser.add_argument('-o', '--outputfile', 
                        help="Optional, the assaysheet for batch sample uploading. If not provided, the file name will be the input template name plus the time stamp.",
                        default=None)
    
    args = parser.parse_args()
    templetfile = args.templetfile
    jsonfile = args.jsonfile
    projectid = args.projectid
    outputfile = args.outputfile
    infositecsvfile = args.infositecsvfile
    #runAuthSampleSubmission(sampledata, sampleType, username, pwd, server)
    parseSamples(templetfile, jsonfile, projectid, outputfile, infositecsvfile)
    return

if __name__ == "__main__":
    main()
